import argparse
import os
import re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

TAG_NAME_PATTERN = re.compile(r"^[A-Za-z_][A-Za-z0-9._:-]*$")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Convert Excel rows to XML using a column-to-tag mapper sheet."
    )
    parser.add_argument("excel_path", help="Path to the Excel file")
    parser.add_argument("output_dir", help="Directory to write XML output")
    parser.add_argument(
        "--multiple-files",
        action="store_true",
        help="Write one XML file per row instead of a single combined XML file",
    )
    parser.add_argument(
        "--root-tag",
        default="root",
        help="Root tag name when writing a combined XML file",
    )
    parser.add_argument(
        "--output-name",
        default="output.xml",
        help="Output file name for the combined XML file",
    )
    return parser.parse_args()


def read_mapper(sheet):
    mapper = {}
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return mapper

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    try:
        column_index = header.index("column_name")
        tag_index = header.index("tag_name")
    except ValueError:
        raise ValueError(
            "Mapper sheet must contain headers 'column_name' and 'tag_name'."
        )

    for row in rows[1:]:
        if row is None:
            continue
        column_name = row[column_index]
        tag_name = row[tag_index]
        if column_name is None or tag_name is None:
            continue
        column_name = str(column_name).strip()
        tag_name = str(tag_name).strip()
        if not TAG_NAME_PATTERN.match(tag_name):
            raise ValueError(
                f"Invalid XML tag name '{tag_name}' in mapper sheet."
            )
        mapper[column_name] = tag_name
    return mapper


def read_data(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows = []
    for row_values in rows[1:]:
        if row_values is None:
            continue
        row_data = {}
        for idx, header_name in enumerate(header):
            if not header_name:
                continue
            row_data[header_name] = row_values[idx] if idx < len(row_values) else None
        data_rows.append(row_data)
    return data_rows


def read_sample_defaults(path):
    if not os.path.exists(path):
        return {}, []

    tree = ET.parse(path)
    root = tree.getroot()
    sample_item = root if root.tag == "item" else root.find("item")
    if sample_item is None:
        return {}, []

    defaults = {}
    order = []
    for child in sample_item:
        defaults[child.tag] = "" if child.text is None else child.text
        order.append(child.tag)
    return defaults, order


def build_xml_element(row_data, mapper, row_tag, sample_defaults, sample_tag_order):
    element = ET.Element(row_tag)
    tag_to_column = {tag_name: column_name for column_name, tag_name in mapper.items()}
    mapped_tags = list(mapper.values())
    ordered_tags = list(sample_tag_order) + [tag for tag in mapped_tags if tag not in sample_tag_order]

    for tag_name in ordered_tags:
        column_name = tag_to_column.get(tag_name)
        value = row_data.get(column_name) if column_name is not None else None
        if value is None:
            value = sample_defaults.get(tag_name, "")
        child = ET.SubElement(element, tag_name)
        child.text = "" if value is None else str(value)
    return element


def write_xml_file(element, path):
    tree = ET.ElementTree(element)
    ET.indent(tree, space="  ", level=0)
    tree.write(path, encoding="utf-8", xml_declaration=True, short_empty_elements=False)


def main():
    args = parse_args()
    os.makedirs(args.output_dir, exist_ok=True)

    workbook = load_workbook(args.excel_path, data_only=True)
    data_sheet_name = "Input"
    mapper_sheet_name = "Mapping"
    if data_sheet_name not in workbook.sheetnames:
        raise ValueError(f"Data sheet '{data_sheet_name}' was not found.")
    if mapper_sheet_name not in workbook.sheetnames:
        raise ValueError(f"Mapper sheet '{mapper_sheet_name}' was not found.")

    data_sheet = workbook[data_sheet_name]
    mapper_sheet = workbook[mapper_sheet_name]

    mapper = read_mapper(mapper_sheet)
    if not mapper:
        raise ValueError("Mapper sheet contains no valid mappings.")

    sample_path = os.path.join(os.path.dirname(os.path.abspath(args.excel_path)), "sample.xml")
    sample_defaults, sample_tag_order = read_sample_defaults(sample_path)

    data_rows = read_data(data_sheet)
    if not data_rows:
        raise ValueError("Data sheet contains no rows to process.")

    if not args.multiple_files:
        root = ET.Element(args.root_tag)
        for row_data in data_rows:
            row_element = build_xml_element(row_data, mapper, "item", sample_defaults, sample_tag_order)
            root.append(row_element)

        output_path = os.path.join(args.output_dir, args.output_name)
        write_xml_file(root, output_path)

        print(f"Wrote combined XML file: {output_path}")
    else:
        for idx, row_data in enumerate(data_rows, start=1):
            row_element = build_xml_element(row_data, mapper, "item", sample_defaults, sample_tag_order)
            filename = f"row_{idx:03}.xml"
            output_path = os.path.join(args.output_dir, filename)
            write_xml_file(row_element, output_path)

        print(f"Wrote {len(data_rows)} XML files to {args.output_dir}")


if __name__ == "__main__":
    main()
